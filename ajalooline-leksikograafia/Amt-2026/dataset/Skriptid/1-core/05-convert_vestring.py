# Created: 2026-06-18 17-20-12
# Author: Madis Jürviste
# Co-Authored-By: Claude Opus 4.8
"""Convert Vestring-1720 xlsx -> Vestring-1720.json (harmonized schema).

Rows with no Estonian headword (col A empty) but carrying an example
(ee näide / saksa tõlge) are continuation examples of the preceding entry —
like Helle's MWUs, they are appended to that entry's mwu[] list rather than
kept as independent, headword-less entries.
"""
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
import importlib
katus_lib = importlib.import_module("00-katus_lib")  # numbered filename requires importlib

SRC = "Katus-ALUSANDMED/new-editions/4-Vestring-1720_Veskimae_20250116_latest.xlsx"
SHEET = "Sheet1"
OUT = "Katus-ALUSANDMED/json-all/Vestring-17XX.json"
SOURCE = "Vestring-17XX"
# A ee vaste B sks vaste C gram jm komm D süno E ee näide F sks tõlge G toim märkus H lk


def ne(v):
    return v is not None and str(v).strip() != ""


def main():
    # worksheet is "unsized" -> load without read_only and walk to max_row
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]
    entries = []
    last = None
    n_attached = n_orphan = 0

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8,
                          values_only=True):
        if not any(ne(c) for c in r):
            continue
        if not ne(r[0]):                       # no headword -> continuation row
            if last is None:                   # no preceding entry: keep standalone
                e = katus_lib.blank_entry(SOURCE)
                e["example-et"] = katus_lib.cell_to_str(r[4])
                e["example-de"] = katus_lib.cell_to_str(r[5])
                e["grammar"] = katus_lib.cell_to_str(r[2])
                e["syn-et"] = katus_lib.cell_to_str(r[3])
                e["comment"] = katus_lib.cell_to_str(r[6])
                e["page"] = katus_lib.cell_to_str(r[7])
                entries.append(e)
                last = e
                n_orphan += 1
                continue
            if ne(r[4]) or ne(r[5]):           # example -> append as MWU item
                last["mwu"].append({
                    "mwu-et": katus_lib.cell_to_str(r[4]),
                    "mwu-de": katus_lib.cell_to_str(r[5]),
                    "page": katus_lib.cell_to_str(r[7]),
                    "comment": katus_lib.cell_to_str(r[6]),
                })
            else:                              # stray row (e.g. only page) -> comment
                bits = [f"{lab}: {katus_lib.cell_to_str(v)}"
                        for lab, v in [("gram", r[2]), ("syn", r[3]),
                                       ("toim", r[6]), ("lk", r[7])] if ne(v)]
                if bits:
                    extra = "stray row -> " + "; ".join(bits)
                    last["comment"] = (extra if last["comment"] == "NULL"
                                       else last["comment"] + "; " + extra)
            n_attached += 1
            continue

        e = katus_lib.blank_entry(SOURCE)
        e["headword-et"] = katus_lib.cell_to_str(r[0])
        e["equiv-de"] = katus_lib.cell_to_str(r[1])
        e["grammar"] = katus_lib.cell_to_str(r[2])    # grammatika jm kommentaar
        e["syn-et"] = katus_lib.cell_to_str(r[3])
        e["example-et"] = katus_lib.cell_to_str(r[4])
        e["example-de"] = katus_lib.cell_to_str(r[5])
        e["comment"] = katus_lib.cell_to_str(r[6])     # toimetaja märkus
        e["page"] = katus_lib.cell_to_str(r[7])
        entries.append(e)
        last = e

    n = katus_lib.dump(entries, OUT, SOURCE)
    total_mwu = sum(len(e["mwu"]) for e in entries)
    print(f"entries: {n}; example rows attached to preceding: {n_attached}; "
          f"orphan standalone: {n_orphan}; total MWU items: {total_mwu}")


if __name__ == "__main__":
    main()
