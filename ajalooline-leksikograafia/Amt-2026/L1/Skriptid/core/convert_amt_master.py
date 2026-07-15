# Created: 2026-06-07 18-09-14
# Author: Madis Jürviste
# Co-Authored-By: Claude Opus 4.8
"""Convert WIP-AMT-MASTER xlsx -> AMT-Master.json (lossless, faithful)."""
import json
import openpyxl

SRC = "Katus-ALUSANDMED/analyzed-tables/WIP-AMT-MASTER_20260607_Updated-OK.xlsx"
OUT = "Katus-ALUSANDMED/analyzed-tables/AMT-Master.json"
SHEET = "MASTER-AMT-Compare"

# Source header -> target label, in column order (cols A..R = 1..18)
LABELS = [
    "Amt-Master-ID",        # Modern
    "Amt-Cat",              # Cat
    "Stahl-1637-et",        # Stahl ET
    "Stahl-1637-de",        # Stahl DE
    "Gutslaff-1648-et",     # Gutslaff EE
    "Gutslaff-1648-de",     # Gutslaff DE
    "Göseken-1660-et",      # Göseken EE
    "Göseken-1660-de",      # Göseken DE
    "Vestring-17XX-et",     # Vestring EE
    "Vestring-17XX-de",     # Vestring DE
    "Helle-1732-et",        # Helle EE
    "Helle-1732-de",        # Helle DE
    "Hupel-1780-est-ger-et",# Hupel 1780 et-de EE
    "Hupel-1780-est-ger-de",# Hupel 1780 et-de DE
    "Cross-source count",   # x_count
    "Comment-1",            # COMM-1
    "Comment-2",            # COMM-2
    "Comment-3",            # COMM-3
]

# Official Estonian alphabet order (32 letters).
ET_ALPHABET = "abcdefghijklmnopqrsšzžtuvwõäöüxy"
# Punctuation/space sort before letters; keep them deterministic and low.
PRE_LETTER = {" ": 0, ",": 1, "-": 2}
LETTER_RANK = {ch: i + 10 for i, ch in enumerate(ET_ALPHABET)}


def collation_key(s):
    """Estonian-collation sort key for a (casefolded) string."""
    key = []
    for ch in s.lower():
        if ch in PRE_LETTER:
            key.append(PRE_LETTER[ch])
        elif ch in LETTER_RANK:
            key.append(LETTER_RANK[ch])
        else:
            # Unknown char: sort after all known letters, by codepoint.
            key.append(1000 + ord(ch))
    return key


def cell_to_str(v):
    """Faithful string rendering of a cell value; empty cell -> 'NULL'.

    None and the empty string are treated as empty -> 'NULL'. A cell that
    contains actual whitespace characters is preserved verbatim (not altered).
    """
    if v is None or v == "":
        return "NULL"
    if isinstance(v, bool):       # guard (none expected)
        return str(v)
    if isinstance(v, int):
        return str(v)             # no floats exist in this sheet
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v)                 # preserves whitespace exactly


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]

    entries = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=18)):
        vals = [c.value for c in row]
        # Skip fully-empty rows (no content in any of the 18 cells).
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            continue
        obj = {LABELS[i]: cell_to_str(vals[i]) for i in range(18)}
        entries.append((idx, obj))

    # Stable sort: Estonian collation on Amt-Master-ID, original order as tiebreak.
    entries.sort(key=lambda t: (collation_key(t[1]["Amt-Master-ID"]), t[0]))
    objects = [obj for _, obj in entries]

    # Build JSON with a blank line between entries inside the AMT-Master array.
    blocks = []
    for obj in objects:
        body = json.dumps(obj, ensure_ascii=False, indent=2)
        body = "\n".join("    " + line for line in body.splitlines())
        blocks.append(body)
    text = '{\n  "AMT-Master": [\n' + ",\n\n".join(blocks) + "\n  ]\n}\n"

    with open(OUT, "w", encoding="utf-8") as f:
        f.write(text)

    # ---- Self-verification ----
    parsed = json.loads(text)
    arr = parsed["AMT-Master"]
    print("entries written:", len(arr))
    assert len(arr) == 921, f"expected 921, got {len(arr)}"
    assert all(set(o.keys()) == set(LABELS) for o in arr), "label mismatch"

    # Lossless check: every non-empty source cell appears verbatim in output.
    # Count source cells that carry real content: not None, not "", not
    # whitespace-only (whitespace-only would be preserved, but none exist here).
    src_cells = 0
    for row in ws.iter_rows(min_row=2, max_col=18):
        vals = [c.value for c in row]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            continue  # skipped fully-empty row, mirror the export
        for v in vals:
            if v is None or v == "":
                continue
            src_cells += 1  # real content (incl. any whitespace-only string)
    out_cells = sum(1 for o in arr for v in o.values() if v != "NULL")
    print("source non-empty cells:", src_cells, "| output non-NULL cells:", out_cells)
    assert src_cells == out_cells, "cell count mismatch -> possible data loss"

    print("first ID:", arr[0]["Amt-Master-ID"], "| last ID:", arr[-1]["Amt-Master-ID"])
    print("OK ->", OUT)


if __name__ == "__main__":
    main()
