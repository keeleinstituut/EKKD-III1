# Created: 2026-06-18 21-46-13
# Author: Madis Jürviste
# Co-Authored-By: Claude Opus 4.8
"""Build an Excel review queue of the still-unresolved Master<->edition links.

One row per Master `<Source>-et` cell that is attested but has an empty
`<Source>-id`. Each row carries the best fuzzy edition suggestion(s) and both
German glosses so a linguist can adjudicate. A DECISION dropdown + free-text
columns make it fillable; a future apply-script can read the result back.
"""
import difflib
import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(__file__))
import link_master as lm

JD = "Katus-ALUSANDMED/json-all"
OUT = "Katus-ALUSANDMED/linkage_review_144.xlsx"
SOURCES = lm.SOURCES


def edition_forms(tag):
    """list of (norm, id, display_form, german); norm-key list for difflib."""
    rows = []
    for e in lm.load(tag):
        pairs = [(e["headword-et"], e["equiv-de"])]
        pairs += [(m["mwu-et"], m["mwu-de"]) for m in e["mwu"]]
        for form, de in pairs:
            if form and form.strip().lower() not in lm.PLACE:
                nk = lm.norm(form)
                if nk:
                    rows.append((nk, e["id"], form, de))
    return rows


def main():
    with open(f"{JD}/AMT-Master_annotated.json", encoding="utf-8") as f:
        master = json.load(f)["AMT-Master"]

    ed = {}
    for tag in SOURCES:
        rows = edition_forms(tag)
        ed[tag] = {"rows": rows, "keys": sorted({r[0] for r in rows})}

    queue = []
    for m in master:
        for tag in SOURCES:
            et = m.get(f"{tag}-et", "NULL")
            if et.strip().lower() in lm.PLACE or m.get(f"{tag}-id"):
                continue
            subs = [lm.norm(s) for s in lm.split_forms(et)]
            subs = [s for s in subs if s and s not in lm.PLACE]
            if not subs:
                continue
            # best candidates over all sub-forms
            cands = []
            for s in subs:
                for c in difflib.get_close_matches(s, ed[tag]["keys"], n=3, cutoff=0.6):
                    sc = round(difflib.SequenceMatcher(None, s, c).ratio(), 3)
                    cands.append((sc, c))
            cands.sort(reverse=True)
            seen, top = set(), []
            for sc, c in cands:
                if c not in seen:
                    seen.add(c)
                    top.append((sc, c))
            best = top[0] if top else None
            sugg_form = sugg_de = sugg_id = ""
            score = ""
            others = ""
            if best:
                sc, c = best
                score = sc
                hits = [r for r in ed[tag]["rows"] if r[0] == c]
                sugg_form = hits[0][2]
                sugg_de = hits[0][3]
                sugg_id = "; ".join(sorted({r[1] for r in hits}))
                others = "; ".join(f"{c2} ({sc2})" for sc2, c2 in top[1:3])
            queue.append({
                "source": tag,
                "master_concept": m["Amt-Master-ID"],
                "master_id": m["id"],
                "master_form": et,
                "master_german": m.get(f"{tag}-de", ""),
                "suggested_edition_form": sugg_form,
                "suggested_edition_german": sugg_de,
                "suggested_edition_id": sugg_id,
                "score": score,
                "other_candidates": others,
            })

    # ---- build workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "review"
    headers = [
        "#", "source", "master_concept", "master_id", "master_form",
        "master_german", "suggested_edition_form", "suggested_edition_german",
        "suggested_edition_id", "score", "other_candidates",
        "DECISION", "correct_edition_id_or_form", "notes",
    ]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="DDDDDD")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    fillable = PatternFill("solid", fgColor="FFF7CC")   # pale yellow for entry cols
    for i, q in enumerate(queue, start=1):
        ws.append([
            i, q["source"], q["master_concept"], q["master_id"], q["master_form"],
            q["master_german"], q["suggested_edition_form"],
            q["suggested_edition_german"], q["suggested_edition_id"], q["score"],
            q["other_candidates"], "", "", "",
        ])
        for col in (12, 13, 14):                        # DECISION / correct / notes
            ws.cell(row=i + 1, column=col).fill = fillable

    # DECISION dropdown
    dv = DataValidation(
        type="list",
        formula1='"accept suggestion,no match in this edition,needs correction"',
        allow_blank=True)
    dv.prompt = "Pick one (or type a correction in the next column)"
    ws.add_data_validation(dv)
    dv.add(f"L2:L{len(queue) + 1}")

    widths = [4, 20, 18, 26, 26, 30, 24, 30, 26, 7, 26, 22, 26, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- instructions sheet ----
    info = wb.create_sheet("instructions", 0)
    lines = [
        ("Katus linkage review — 144 unresolved Master↔edition cells", True),
        ("", False),
        ("Each row in the 'review' sheet is one AMT-Master source cell whose "
         "historical Estonian form could NOT be linked automatically.", False),
        ("Auto-linking already covered ~92%; these are the spelling-divergent "
         "or possibly-absent remainder.", False),
        ("", False),
        ("Columns A–K are read-only context (don't edit). Fill the yellow "
         "columns L–N:", True),
        ("  L  DECISION  — dropdown: 'accept suggestion' / 'no match in this "
         "edition' / 'needs correction'", False),
        ("  M  correct_edition_id_or_form — if 'needs correction', paste the "
         "right edition id (e.g. he-019...) or the correct edition headword", False),
        ("  N  notes — anything useful (sense split, MWU, dialect, etc.)", False),
        ("", False),
        ("Context columns:", True),
        ("  master_form / master_german  = how AMT-Master records this source's "
         "form and gloss", False),
        ("  suggested_edition_form / _german / _id = best fuzzy candidate in the "
         "edition + its German gloss (compare the GERMAN to confirm same word)", False),
        ("  score = string similarity 0–1 (higher = more similar spelling)", False),
        ("  other_candidates = next-best edition forms, for reference", False),
        ("", False),
        ("Tip: the German gloss is the safest confirmation — if the suggested "
         "edition's German matches master_german, it's almost certainly the same "
         "word despite different spelling.", False),
        ("Once filled, send it back and a script will write the confirmed links "
         "into both the Master (<Source>-id) and the editions (master-id).", False),
    ]
    for r, (text, bold) in enumerate(lines, start=1):
        cell = info.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 110

    wb.save(OUT)
    print(f"wrote {len(queue)} review rows -> {OUT}")
    # quick per-source tally
    from collections import Counter
    tally = Counter(q["source"] for q in queue)
    for tag in SOURCES:
        print(f"  {tag:22} {tally[tag]}")


if __name__ == "__main__":
    main()
